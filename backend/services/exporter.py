"""Excel model and PDF tearsheet generation.

The Excel workbook is a real model, not a data dump: assumption cells
are conventional blue inputs and every output is a live formula
(=SUM, discount factors, Gordon terminal value, IRR), so it recalculates
when opened and edited in Excel.
"""

import io
import textwrap
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdf_canvas

from services import market_data

GOLD = "D4AF37"
INK = "1A1A1F"
INPUT_BLUE = "1F4E9C"

_TITLE = Font(name="Calibri", size=16, bold=True, color=INK)
_SUBTITLE = Font(name="Calibri", size=11, color="6B6B76")
_SECTION = Font(name="Calibri", size=11, bold=True, color=INK)
_LABEL = Font(name="Calibri", size=11, color=INK)
_INPUT = Font(name="Calibri", size=11, color=INPUT_BLUE)
_OUTPUT_BOLD = Font(name="Calibri", size=11, bold=True, color=INK)
_SECTION_BORDER = Border(bottom=Side(style="thin", color=GOLD))
_GOLD_FILL = PatternFill("solid", fgColor="FBF3DB")

_MM = '#,##0,,'  # raw dollars rendered in millions
_PCT = "0.0%"
_PRICE = "#,##0.00"
_MULT = '0.00"×"'


def _section(ws, row: int, text: str) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION
    cell.border = _SECTION_BORDER
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = _SECTION_BORDER


def _kv(ws, row: int, label: str, value: Any, fmt: str, font: Font = _INPUT) -> None:
    ws.cell(row=row, column=1, value=label).font = _LABEL
    cell = ws.cell(row=row, column=2, value=value)
    cell.font = font
    cell.number_format = fmt


def _titles(ws, model_name: str, name: str, symbol: str) -> None:
    ws.cell(row=1, column=1, value="AURUM").font = Font(
        name="Calibri", size=12, bold=True, color=GOLD
    )
    ws.cell(row=2, column=1, value=model_name).font = _TITLE
    ws.cell(row=3, column=1, value=f"{name} ({symbol}) · generated {date.today().isoformat()} · aurum-terminal.vercel.app").font = _SUBTITLE
    ws.column_dimensions["A"].width = 30
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _dcf_sheet(ws, snapshot: dict, dcf: dict) -> None:
    _titles(ws, "Discounted Cash Flow Model", snapshot["name"], snapshot["ticker"])

    _section(ws, 5, "ASSUMPTIONS (edit the blue cells)")
    _kv(ws, 6, "Base free cash flow (TTM)", dcf["base_fcf"], _MM)
    _kv(ws, 7, "FCF growth, years 1-5", dcf["suggested_growth"] / 100, _PCT)
    _kv(ws, 8, "Terminal growth", dcf["suggested_terminal"] / 100, _PCT)
    _kv(ws, 9, "Discount rate (WACC)", dcf["suggested_discount"] / 100, _PCT)
    _kv(ws, 10, "Net debt", dcf["net_debt"], _MM)
    _kv(ws, 11, "Shares outstanding", dcf["shares_outstanding"], "#,##0,,")
    _kv(ws, 12, "Current share price", snapshot["price"], _PRICE)

    _section(ws, 14, "PROJECTION (US$ in millions)")
    ws.cell(row=15, column=1, value="Year").font = _LABEL
    for i in range(5):
        col = 2 + i
        year_cell = ws.cell(row=15, column=col, value=i + 1)
        year_cell.font = _OUTPUT_BOLD
        year_cell.alignment = Alignment(horizontal="right")

    ws.cell(row=16, column=1, value="Free cash flow").font = _LABEL
    ws.cell(row=16, column=2, value="=$B$6*(1+$B$7)")
    for i in range(1, 5):
        prev = get_column_letter(1 + i)
        ws.cell(row=16, column=2 + i, value=f"={prev}16*(1+$B$7)")

    ws.cell(row=17, column=1, value="Discount factor").font = _LABEL
    for i in range(5):
        col_letter = get_column_letter(2 + i)
        ws.cell(row=17, column=2 + i, value=f"=1/(1+$B$9)^{col_letter}15")

    ws.cell(row=18, column=1, value="PV of free cash flow").font = _LABEL
    for i in range(5):
        col_letter = get_column_letter(2 + i)
        ws.cell(row=18, column=2 + i, value=f"={col_letter}16*{col_letter}17")

    for row, fmt in ((16, _MM), (17, "0.000"), (18, _MM)):
        for i in range(5):
            ws.cell(row=row, column=2 + i).number_format = fmt

    _section(ws, 20, "VALUATION OUTPUT")
    outputs = [
        (21, "Sum of PV, years 1-5", "=SUM(B18:F18)", _MM, _LABEL),
        (22, "Terminal value (Gordon)", "=F16*(1+$B$8)/($B$9-$B$8)", _MM, _LABEL),
        (23, "PV of terminal value", "=B22/(1+$B$9)^5", _MM, _LABEL),
        (24, "Enterprise value", "=B21+B23", _MM, _OUTPUT_BOLD),
        (25, "Less: net debt", "=-B10", _MM, _LABEL),
        (26, "Equity value", "=B24+B25", _MM, _OUTPUT_BOLD),
        (27, "Implied value per share", "=B26/B11", _PRICE, _OUTPUT_BOLD),
        (28, "Current share price", "=B12", _PRICE, _LABEL),
        (29, "Implied upside / (downside)", "=B27/B28-1", _PCT, _OUTPUT_BOLD),
    ]
    for row, label, formula, fmt, font in outputs:
        _kv(ws, row, label, formula, fmt, font)
    ws.cell(row=27, column=2).fill = _GOLD_FILL
    ws.cell(row=29, column=2).fill = _GOLD_FILL

    ws.cell(
        row=31,
        column=1,
        value="Simplified 5-year DCF for educational exploration. Not investment advice.",
    ).font = _SUBTITLE


def _lbo_sheet(ws, snapshot: dict, lbo: dict) -> None:
    _titles(ws, "Leveraged Buyout Model", snapshot["name"], snapshot["ticker"])

    _section(ws, 5, "ASSUMPTIONS (edit the blue cells)")
    _kv(ws, 6, "EBITDA (TTM)", lbo["ebitda"], _MM)
    _kv(ws, 7, "Entry multiple (EV/EBITDA)", lbo["suggested_entry_multiple"], _MULT)
    _kv(ws, 8, "Debt financing (% of EV)", 0.6, _PCT)
    _kv(ws, 9, "Interest rate", 0.08, _PCT)
    _kv(ws, 10, "EBITDA growth", 0.06, _PCT)
    _kv(ws, 11, "Exit multiple (EV/EBITDA)", lbo["suggested_entry_multiple"], _MULT)

    _section(ws, 13, "SOURCES & USES")
    _kv(ws, 14, "Entry enterprise value", "=B6*B7", _MM, _LABEL)
    _kv(ws, 15, "Entry debt", "=B14*B8", _MM, _LABEL)
    _kv(ws, 16, "Sponsor equity check", "=B14-B15", _MM, _OUTPUT_BOLD)

    _section(ws, 18, "PROJECTION (US$ in millions, 50% cash sweep)")
    ws.cell(row=19, column=1, value="Year").font = _LABEL
    for i in range(5):
        cell = ws.cell(row=19, column=2 + i, value=i + 1)
        cell.font = _OUTPUT_BOLD
        cell.alignment = Alignment(horizontal="right")

    ws.cell(row=20, column=1, value="EBITDA").font = _LABEL
    ws.cell(row=20, column=2, value="=$B$6*(1+$B$10)")
    for i in range(1, 5):
        prev = get_column_letter(1 + i)
        ws.cell(row=20, column=2 + i, value=f"={prev}20*(1+$B$10)")

    ws.cell(row=21, column=1, value="Interest expense").font = _LABEL
    ws.cell(row=21, column=2, value="=$B$15*$B$9")
    for i in range(1, 5):
        prev = get_column_letter(1 + i)
        ws.cell(row=21, column=2 + i, value=f"={prev}23*$B$9")

    ws.cell(row=22, column=1, value="Debt paydown (50% sweep)").font = _LABEL
    for i in range(5):
        col_letter = get_column_letter(2 + i)
        ws.cell(row=22, column=2 + i, value=f"=MAX(0,({col_letter}20-{col_letter}21)*0.5)")

    ws.cell(row=23, column=1, value="Debt, end of year").font = _LABEL
    ws.cell(row=23, column=2, value="=MAX(0,$B$15-B22)")
    for i in range(1, 5):
        prev = get_column_letter(1 + i)
        col_letter = get_column_letter(2 + i)
        ws.cell(row=23, column=2 + i, value=f"=MAX(0,{prev}23-{col_letter}22)")

    for row in (20, 21, 22, 23):
        for i in range(5):
            ws.cell(row=row, column=2 + i).number_format = _MM

    _section(ws, 25, "RETURNS")
    _kv(ws, 26, "Exit enterprise value", "=F20*B11", _MM, _LABEL)
    _kv(ws, 27, "Exit equity value", "=B26-F23", _MM, _OUTPUT_BOLD)
    _kv(ws, 28, "Multiple on invested capital", "=B27/B16", _MULT, _OUTPUT_BOLD)
    _kv(ws, 29, "IRR (5-year hold)", "=(B27/B16)^(1/5)-1", _PCT, _OUTPUT_BOLD)
    ws.cell(row=28, column=2).fill = _GOLD_FILL
    ws.cell(row=29, column=2).fill = _GOLD_FILL

    ws.cell(
        row=31,
        column=1,
        value="Simplified paper-LBO for educational exploration. Not investment advice.",
    ).font = _SUBTITLE


def _financials_sheet(ws, snapshot: dict) -> None:
    _titles(ws, "Financial Statements", snapshot["name"], snapshot["ticker"])
    row = 5
    for statement, heading in (
        ("income", "INCOME STATEMENT (US$ in millions)"),
        ("balance", "BALANCE SHEET (US$ in millions)"),
        ("cash", "CASH FLOW STATEMENT (US$ in millions)"),
    ):
        data = market_data.get_financials(snapshot["ticker"], statement)
        if not data:
            continue
        _section(ws, row, heading)
        row += 1
        for i, period in enumerate(data["periods"]):
            cell = ws.cell(row=row, column=2 + i, value=f"FY {period}")
            cell.font = _OUTPUT_BOLD
            cell.alignment = Alignment(horizontal="right")
        row += 1
        for line in data["rows"]:
            ws.cell(row=row, column=1, value=line["label"]).font = _LABEL
            for i, value in enumerate(line["values"]):
                cell = ws.cell(row=row, column=2 + i, value=value)
                cell.number_format = _MM
            row += 1
        row += 1


def build_excel_model(raw_ticker: str) -> tuple[bytes, str] | None:
    """Build the workbook; returns (bytes, filename) or None if unknown."""
    snapshot = market_data.get_company_snapshot(raw_ticker)
    if snapshot is None:
        return None
    dcf = market_data.get_dcf_inputs(snapshot["ticker"])
    lbo = market_data.get_lbo_inputs(snapshot["ticker"])

    workbook = Workbook()
    workbook.remove(workbook.active)

    if dcf and dcf.get("base_fcf") and dcf["base_fcf"] > 0 and dcf.get("shares_outstanding"):
        _dcf_sheet(workbook.create_sheet("DCF Model"), snapshot, dcf)
    if lbo and lbo.get("ebitda") and lbo["ebitda"] > 0:
        _lbo_sheet(workbook.create_sheet("LBO Model"), snapshot, lbo)
    _financials_sheet(workbook.create_sheet("Financials"), snapshot)

    if not workbook.sheetnames:
        return None
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), f"{snapshot['ticker']}-aurum-model.xlsx"


# ---------------------------------------------------------------------------
# PDF tearsheet
# ---------------------------------------------------------------------------

_PDF_GOLD = HexColor("#B08E2E")
_PDF_INK = HexColor("#1A1A1F")
_PDF_GRAY = HexColor("#6B6B76")
_PDF_GAIN = HexColor("#0E9F6E")
_PDF_LOSS = HexColor("#D6293A")


def build_pdf_tearsheet(raw_ticker: str) -> tuple[bytes, str] | None:
    """One-page company tearsheet; returns (bytes, filename) or None."""
    snapshot = market_data.get_company_snapshot(raw_ticker)
    if snapshot is None:
        return None
    symbol = snapshot["ticker"]
    stats = market_data.get_key_stats(symbol) or {}
    dcf = market_data.get_dcf_inputs(symbol)

    buffer = io.BytesIO()
    page = pdf_canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 54
    y = height - 56

    def fmt_big(value: Any) -> str:
        if not isinstance(value, (int, float)):
            return "—"
        for cutoff, suffix in ((1e12, "T"), (1e9, "B"), (1e6, "M")):
            if abs(value) >= cutoff:
                return f"{value / cutoff:,.2f}{suffix}"
        return f"{value:,.0f}"

    # Header band
    page.setFont("Helvetica-Bold", 11)
    page.setFillColor(_PDF_GOLD)
    page.drawString(margin, y, "AURUM")
    page.setFont("Helvetica", 9)
    page.setFillColor(_PDF_GRAY)
    page.drawString(margin + 52, y, "EQUITY TEARSHEET")
    page.drawRightString(width - margin, y, date.today().strftime("%B %d, %Y"))
    y -= 26

    page.setFont("Helvetica-Bold", 20)
    page.setFillColor(_PDF_INK)
    page.drawString(margin, y, snapshot["name"][:58])
    y -= 16
    page.setFont("Helvetica", 10)
    page.setFillColor(_PDF_GRAY)
    exchange = snapshot.get("exchange") or ""
    page.drawString(margin, y, f"{symbol} · {exchange} · {snapshot['currency']}")

    change = snapshot["change"]
    page.setFont("Helvetica-Bold", 18)
    page.setFillColor(_PDF_INK)
    page.drawRightString(width - margin, y + 14, f"{snapshot['price']:,.2f}")
    page.setFont("Helvetica-Bold", 10)
    page.setFillColor(_PDF_GAIN if change >= 0 else _PDF_LOSS)
    sign = "+" if change >= 0 else ""
    page.drawRightString(
        width - margin, y, f"{sign}{change:,.2f} ({sign}{snapshot['change_percent']:.2f}%)"
    )
    y -= 14

    page.setStrokeColor(_PDF_GOLD)
    page.setLineWidth(1.4)
    page.line(margin, y, width - margin, y)
    y -= 24

    # Key statistics grid
    page.setFont("Helvetica-Bold", 10)
    page.setFillColor(_PDF_INK)
    page.drawString(margin, y, "KEY STATISTICS")
    y -= 16
    week52 = (
        f"{stats.get('week52_low'):,.2f} – {stats.get('week52_high'):,.2f}"
        if isinstance(stats.get("week52_low"), (int, float))
        and isinstance(stats.get("week52_high"), (int, float))
        else "—"
    )
    fmt_num = lambda v: f"{v:,.2f}" if isinstance(v, (int, float)) else "—"  # noqa: E731
    fmt_pct = lambda v: f"{v:.2f}%" if isinstance(v, (int, float)) else "—"  # noqa: E731
    entries = [
        ("Market cap", fmt_big(snapshot.get("market_cap"))),
        ("Trailing P/E", fmt_num(stats.get("trailing_pe"))),
        ("Forward P/E", fmt_num(stats.get("forward_pe"))),
        ("EPS (TTM)", fmt_num(stats.get("eps"))),
        ("Dividend yield", fmt_pct(stats.get("dividend_yield"))),
        ("Beta", fmt_num(stats.get("beta"))),
        ("52-week range", week52),
        ("Avg volume", fmt_big(stats.get("avg_volume"))),
        ("Revenue (TTM)", fmt_big(stats.get("revenue"))),
        ("Profit margin", fmt_pct(stats.get("profit_margin"))),
        ("Return on equity", fmt_pct(stats.get("return_on_equity"))),
        ("Free cash flow", fmt_big(stats.get("free_cash_flow"))),
    ]
    col_width = (width - 2 * margin) / 3
    for i, (label, value) in enumerate(entries):
        cx = margin + (i % 3) * col_width
        cy = y - (i // 3) * 30
        page.setFont("Helvetica", 7.5)
        page.setFillColor(_PDF_GRAY)
        page.drawString(cx, cy, label.upper())
        page.setFont("Helvetica-Bold", 11)
        page.setFillColor(_PDF_INK)
        page.drawString(cx, cy - 12, value)
    y -= 30 * 4 + 8

    # DCF summary box
    if dcf and dcf.get("base_fcf") and dcf["base_fcf"] > 0 and dcf.get("shares_outstanding"):
        growth = dcf["suggested_growth"] / 100
        terminal = dcf["suggested_terminal"] / 100
        rate = dcf["suggested_discount"] / 100
        fcf = dcf["base_fcf"]
        pv_sum = 0.0
        for year in range(1, 6):
            fcf_year = dcf["base_fcf"] * (1 + growth) ** year
            pv_sum += fcf_year / (1 + rate) ** year
            fcf = fcf_year
        terminal_value = fcf * (1 + terminal) / (rate - terminal)
        enterprise = pv_sum + terminal_value / (1 + rate) ** 5
        equity = enterprise - dcf["net_debt"]
        per_share = equity / dcf["shares_outstanding"]
        upside = (per_share / snapshot["price"] - 1) * 100

        box_height = 64
        page.setFillColor(HexColor("#FBF6E7"))
        page.setStrokeColor(_PDF_GOLD)
        page.roundRect(margin, y - box_height, width - 2 * margin, box_height, 6, fill=1)
        page.setFont("Helvetica-Bold", 10)
        page.setFillColor(_PDF_INK)
        page.drawString(margin + 14, y - 18, "DCF SNAPSHOT (BASE ASSUMPTIONS)")
        page.setFont("Helvetica", 8.5)
        page.setFillColor(_PDF_GRAY)
        page.drawString(
            margin + 14,
            y - 32,
            f"{dcf['suggested_growth']:.1f}% FCF growth · {dcf['suggested_terminal']:.1f}% terminal · {dcf['suggested_discount']:.1f}% discount rate · 5-year horizon",
        )
        page.setFont("Helvetica-Bold", 13)
        page.setFillColor(_PDF_INK)
        page.drawString(margin + 14, y - 50, f"Implied value: {per_share:,.2f}")
        page.setFillColor(_PDF_GAIN if upside >= 0 else _PDF_LOSS)
        page.drawString(
            margin + 190, y - 50, f"{'+' if upside >= 0 else ''}{upside:.1f}% vs. market"
        )
        y -= box_height + 22

    # Company profile
    if stats.get("summary"):
        page.setFont("Helvetica-Bold", 10)
        page.setFillColor(_PDF_INK)
        page.drawString(margin, y, "BUSINESS")
        y -= 14
        meta = " · ".join(
            str(part)
            for part in (stats.get("sector"), stats.get("industry"))
            if part
        )
        if meta:
            page.setFont("Helvetica", 8.5)
            page.setFillColor(_PDF_GRAY)
            page.drawString(margin, y, meta)
            y -= 14
        page.setFont("Helvetica", 9)
        page.setFillColor(_PDF_INK)
        for line in textwrap.wrap(stats["summary"], width=104)[:6]:
            page.drawString(margin, y, line)
            y -= 12

    # Footer
    page.setFont("Helvetica", 7.5)
    page.setFillColor(_PDF_GRAY)
    page.drawString(
        margin,
        44,
        "Generated by Aurum · aurum-terminal.vercel.app · Built by Jeriel De Leon · Data via Yahoo Finance",
    )
    page.drawString(margin, 33, "Educational analysis, not investment advice.")

    page.showPage()
    page.save()
    return buffer.getvalue(), f"{symbol}-aurum-tearsheet.pdf"
